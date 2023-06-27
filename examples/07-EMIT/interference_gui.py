import sys
from PySide2 import QtWidgets, QtUiTools, QtGui
from pyaedt.emit_core.emit_constants import InterfererType, ResultType, TxRxMode
from pyaedt import Emit
import pyaedt
from openpyxl.styles import PatternFill
import openpyxl
import os
import pyaedt.generic.constants as consts
from interference_classification import interference_classification

# Check that emit is a compatible version
emitapp_desktop_version = "2023.2"
if emitapp_desktop_version < "2023.2":
    print("Must have v2023.2 or later")
    sys.exit()

# Launch emit
non_graphical = False
new_thread = True
d = pyaedt.launch_desktop(emitapp_desktop_version, non_graphical, new_thread)
emitapp = Emit(pyaedt.generate_unique_project_name())

# Add emitapi to system path
emit_path = emitapp._desktop_install_dir + "/Delcross"
sys.path.append(emit_path)
import EmitApiPython
api = EmitApiPython.EmitApi()

# Define .ui file for GUI
Ui_MainWindow, _ = QtUiTools.loadUiType("gui.ui")

class DoubleDelegate(QtWidgets.QStyledItemDelegate):
    def __init__(self, decimals):
        super().__init__()
        self.decimals = decimals

    def createEditor(self, parent, option, index):
        editor = super().createEditor(parent, option, index)
        if isinstance(editor, QtWidgets.QLineEdit):
            validator = QtGui.QDoubleValidator(parent)
            validator.setDecimals(self.decimals)
            editor.setValidator(validator)
        return editor

class MainWindow(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.setupUi(self)
        self.setup_widgets()
    
    def setup_widgets(self): # Widget definitions

        # Widget definitions for file selection/tab management
        self.file_select_btn = self.findChild(QtWidgets.QToolButton, "file_select_btn")
        self.file_path_box = self.findChild(QtWidgets.QLineEdit, "file_path_box") 
        self.design_name_dropdown = self.findChild(QtWidgets.QComboBox, "design_name_dropdown")
        self.tab_widget = self.findChild(QtWidgets.QTabWidget, "tab_widget")

        # Widget definitions for protection level classification
        self.protection_results_btn = self.findChild(QtWidgets.QPushButton, "protection_results_btn")
        self.protection_matrix = self.findChild(QtWidgets.QTableWidget, "protection_matrix")    
        self.protection_legend_table = self.findChild(QtWidgets.QTableWidget, "protection_legend_table")
        self.damage_check = self.findChild(QtWidgets.QCheckBox, "damage_check")
        self.overload_check = self.findChild(QtWidgets.QCheckBox, "overload_check")
        self.intermodulation_check = self.findChild(QtWidgets.QCheckBox, "intermodulation_check")
        self.desensitization_check = self.findChild(QtWidgets.QCheckBox, "desensitization_check")
        self.protection_export_btn = self.findChild(QtWidgets.QPushButton, "protection_export_btn")
        self.radio_specific_levels = self.findChild(QtWidgets.QCheckBox, "radio_specific_levels")
        self.radio_dropdown = self.findChild(QtWidgets.QComboBox, "radio_dropdown")
        self.protection_save_img_btn = self.findChild(QtWidgets.QPushButton, 'protection_save_img_btn')

        # Setup for protection level buttons and table
        self.protection_export_btn.setEnabled(False)
        self.protection_save_img_btn.setEnabled(False)
        self.file_select_btn.clicked.connect(self.open_file_dialog) 
        self.protection_export_btn.clicked.connect(self.save_results_excel)
        self.protection_results_btn.clicked.connect(self.protection_results)
        self.protection_legend_table.resizeRowsToContents()
        self.protection_legend_table.resizeColumnsToContents()
        self.damage_check.stateChanged.connect(self.protection_results)
        self.overload_check.stateChanged.connect(self.protection_results)
        self.intermodulation_check.stateChanged.connect(self.protection_results)
        self.desensitization_check.stateChanged.connect(self.protection_results)
        self.protection_legend_table.setEditTriggers(QtWidgets.QTableWidget.DoubleClicked)
        self.global_protection_level = True
        self.protection_levels = {}
        values = [float(self.protection_legend_table.item(row, 0).text()) for row in range(self.protection_legend_table.rowCount())]
        self.protection_levels['Global'] = values
        self.changing = False
        self.radio_dropdown.currentIndexChanged.connect(self.radio_dropdown_changed)
        self.protection_legend_table.itemChanged.connect(self.table_changed)
        self.protection_save_img_btn.clicked.connect(self.save_image)

        # Widget definitions for interference type
        self.interference_results_btn = self.findChild(QtWidgets.QPushButton, "interference_results_btn")
        self.interference_matrix = self.findChild(QtWidgets.QTableWidget, "interference_matrix")    
        self.interference_legend_table = self.findChild(QtWidgets.QTableWidget, "interference_legend_table")
        self.in_in_check = self.findChild(QtWidgets.QCheckBox, "in_in_check")
        self.in_out_check = self.findChild(QtWidgets.QCheckBox, "in_out_check")
        self.out_in_check = self.findChild(QtWidgets.QCheckBox, "out_in_check")
        self.out_out_check = self.findChild(QtWidgets.QCheckBox, "out_out_check")
        self.interference_export_btn = self.findChild(QtWidgets.QPushButton, "interference_export_btn")
        self.interference_save_img_btn = self.findChild(QtWidgets.QPushButton, 'interference_save_img_btn')
        
        # Setup for interference type buttons and table
        self.interference_export_btn.setEnabled(False)
        self.interference_save_img_btn.setEnabled(False)
        self.interference_export_btn.clicked.connect(self.save_results_excel)
        self.interference_results_btn.clicked.connect(self.interference_results)
        self.interference_legend_table.resizeRowsToContents() 
        self.interference_legend_table.resizeColumnsToContents()
        self.in_in_check.stateChanged.connect(self.interference_results)
        self.in_out_check.stateChanged.connect(self.interference_results)
        self.out_in_check.stateChanged.connect(self.interference_results)
        self.out_out_check.stateChanged.connect(self.interference_results)
        self.radio_specific_levels.stateChanged.connect(self.radio_specific)
        self.interference_save_img_btn.clicked.connect(self.save_image)

        # Color definition dictionary and previous project/design names
        self.color_dict = {"green": [QtGui.QColor(125, 115, 202),'#7d73ca'], 
                           "yellow":[QtGui.QColor(211, 89, 162), '#d359a2'], 
                           "orange": [QtGui.QColor(255, 99, 97), '#ff6361'], 
                           "red": [QtGui.QColor(255, 166, 0), '#ffa600'], 
                           "white": [QtGui.QColor("white"),'#ffffff']}
        self.previous_design = ''
        self.previous_project = ''

        # Set the legend tables to strech resize mode
        header = self.protection_legend_table.horizontalHeader()
        v_header = self.protection_legend_table.verticalHeader()

        header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeMode.Stretch)
        v_header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeMode.Stretch)
        v_header.setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeMode.Stretch)
        v_header.setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeMode.Stretch)
        v_header.setSectionResizeMode(3, QtWidgets.QHeaderView.ResizeMode.Stretch)

        header = self.interference_legend_table.horizontalHeader()
        v_header = self.interference_legend_table.verticalHeader()

        header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeMode.Stretch)
        v_header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeMode.Stretch)
        v_header.setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeMode.Stretch)
        v_header.setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeMode.Stretch)
        v_header.setSectionResizeMode(3, QtWidgets.QHeaderView.ResizeMode.Stretch)

        # Input validation for protection level legend table
        delegate = DoubleDelegate(decimals=2)
        self.protection_legend_table.setItemDelegateForColumn(0, delegate)

    def open_file_dialog(self):
        'Open the file dialog and open the corresponging EMIT Project/Design'
        fname = QtWidgets.QFileDialog.getOpenFileName(self, "Select EMIT Project", "", "Ansys Electronics Desktop Files (*.aedt)", )
        if fname: 
            self.file_path_box.setText(fname[0])

            # Close previous project and open specified one
            emitapp.close_project()       
            emitapp.load_project(self.file_path_box.text())
            
            # Populate design dropdown with all design names
            designs = emitapp.design_list
            self.design_name_dropdown.clear()
            for i in designs:
                self.design_name_dropdown.addItem(i)
                
            self.design_name_dropdown.setEnabled(True)

            if self.radio_specific_levels.isEnabled():
                self.radio_specific_levels.setChecked(False)
                self.radio_dropdown.clear()
                self.radio_dropdown.setEnabled(False)
                self.protection_levels = {}
                values = [float(self.protection_legend_table.item(row, 0).text()) for row in range(self.protection_legend_table.rowCount())]
                self.protection_levels['Global'] = values

            self.radio_specific_levels.setEnabled(True)
    
    def radio_specific(self):
        'Enable radio specific proteciton levels'
        self.radio_dropdown.setEnabled(self.radio_specific_levels.isChecked())
        self.radio_dropdown.clear()
        if self.radio_dropdown.isEnabled():
            emitapp.set_active_design(self.design_name_dropdown.currentText())
            radios = emitapp.modeler.components.get_radios()
            values = [float(self.protection_legend_table.item(row, 0).text()) for row in range(self.protection_legend_table.rowCount())]
            for radio in radios:
                if radios[radio].has_rx_channels():
                    self.protection_levels[radio] = values
                    self.radio_dropdown.addItem(radio)
        else:
            self.radio_dropdown.clear()
            values = [float(self.protection_legend_table.item(row, 0).text()) for row in range(self.protection_legend_table.rowCount())]
            self.protection_levels['Global'] = values
        self.global_protection_level = not self.radio_specific_levels.isChecked()
    
    def radio_dropdown_changed(self):
        'Update table when radio changes'
        if self.radio_dropdown.isEnabled():
            self.changing = True
            for row in range(self.protection_legend_table.rowCount()):
                item = self.protection_legend_table.item(row, 0)
                item.setText(str(self.protection_levels[self.radio_dropdown.currentText()][row]))
            self.changing = False

    def table_changed(self):
        'Save new table values'
        if self.changing == False:
            values = [float(self.protection_legend_table.item(row, 0).text()) for row in range(self.protection_legend_table.rowCount())]
            if self.radio_dropdown.currentText() == '':
                index = 'Global'
            else:
                index = self.radio_dropdown.currentText()
            self.protection_levels[index] = values

    def save_image(self):
        'Save scenario matrix as png'
        if self.tab_widget.currentIndex() == 0:
            table = self.protection_matrix
        else:
            table = self.interference_matrix

        fname = QtWidgets.QFileDialog.getSaveFileName(self, "Save Scenario Matrix", "Scenario Matrix", "png (*.png)")
        image = QtGui.QImage(table.size(), QtGui.QImage.Format_ARGB32)
        table.render(image)
        image.save(fname[0])

    def save_results_excel(self):
        ' Save interference/protection matrix as an excel file'
        fname = QtWidgets.QFileDialog.getSaveFileName(self, "Save Scenario Matrix", "Protection Level Classification", "xlsx (*.xlsx)")
        
        if self.tab_widget.currentIndex() == 0:
            table = self.protection_matrix
        else:
            table = self.interference_matrix

        if fname:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            header = self.tx_radios
            header.insert(0, "Tx/Rx")
            worksheet.append(header)
            for row in range(2, table.rowCount()+2):
                worksheet.cell(row = row, column = 1, value = str(self.rx_radios[row-2]))
                for col in range(2, table.columnCount()+2):
                    text = str(table.item(row-2, col-2).text())
                    worksheet.cell(row = row, column = col, value = text)
                    cell = worksheet.cell(row, col)
                    cell.fill = PatternFill(start_color = self.color_dict[self.all_colors[col-2][row-2]][1][1:], 
                                            end_color = self.color_dict[self.all_colors[col-2][row-2]][1][1:], 
                                            fill_type = "solid")
            workbook.save(fname[0])

    def interference_results(self):
        ' Run simulation to classify results according to interference type'

        # Initialize filter check marks and expected filter results
        self.interference_checks = [self.in_in_check.isChecked(), self.out_in_check.isChecked(),
                                  self.in_out_check.isChecked(), self.out_out_check.isChecked()]

        self.interference_filters =["TxFundamental:In-band", ["TxHarmonic/Spurious:In-band","Intermod:In-band", "Broadband:In-band"], 
                                    "TxFundamental:Out-of-band", ["TxHarmonic/Spurious:Out-of-band","Intermod:Out-of-band", "Broadband:Out-of-band"]]

        # Create list of problem types to analyze according to inputted filters
        filter = [i for (i,v) in zip(self.interference_filters, self.interference_checks) if v]

        if self.file_path_box.text() != "" and self.design_name_dropdown.currentText() != "":
            if self.previous_design != self.design_name_dropdown.currentText() or self.previous_project != self.file_path_box.text():
                # emitapp.close_project()
                self.previous_design = self.design_name_dropdown.currentText()
                self.previous_project = self.file_path_box.text()
                emitapp.set_active_design(self.design_name_dropdown.currentText())
            
                # Get results and radios
                self.rev = emitapp.results.analyze()
                self.tx_interferer = InterfererType().TRANSMITTERS
                self.rx_radios = self.rev.get_receiver_names()
                self.tx_radios = self.rev.get_interferer_names(self.tx_interferer)

                # Check if design is valid
                if self.tx_radios is None or self.rx_radios is None:
                    return

            # Classify the interference
            # ~~~~~~~~~~~~~~~~~~~~~~~~~~~
            # Iterate over all the transmitters and receivers and compute the power
            # at the input to each receiver due to each of the transmitters. Computes
            # which, if any, type of interference occured.
            self.all_colors, self.power_matrix = interference_classification(emitapp, use_filter = True, filter = filter)

            # Save project and plot results on table widget
            emitapp.save_project()
            self.populate_table()

    def protection_results(self):

        if self.global_protection_level:
            damage_threshold = self.protection_levels['Global'][0]
            overload_threshold = self.protection_levels['Global'][1]
            intermod_threshold = self.protection_levels['Global'][2]
            desense_threshold = self.protection_levels['Global'][3]

        self.protection_checks = [self.damage_check.isChecked(), self.overload_check.isChecked(),
                                  self.intermodulation_check.isChecked(), self.desensitization_check.isChecked()]

        self.protection_filters = ['damage', 'overload', 'intermodulation', 'desensitization']

        filter = [i for (i,v) in zip(self.protection_filters, self.protection_checks) if v]

        if self.file_path_box.text() != "" and self.design_name_dropdown.currentText() != "":
            if self.previous_design != self.design_name_dropdown.currentText() or self.previous_project != self.file_path_box.text():
                self.previous_design = self.design_name_dropdown.currentText()
                self.previous_project = self.file_path_box.text()
                emitapp.set_active_design(self.design_name_dropdown.currentText())

                # Get results and design radios
                self.modeRx = TxRxMode().RX
                self.modeTx = TxRxMode().TX
                self.mode_power = ResultType().POWER_AT_RX
                self.tx_interferer = InterfererType().TRANSMITTERS
                self.rev = emitapp.results.analyze()
                self.rx_radios = self.rev.get_receiver_names()
                self.tx_radios = self.rev.get_interferer_names(self.tx_interferer)
                self.domain = emitapp.results.interaction_domain()
                self.radios = emitapp.modeler.components.get_radios()

            if self.tx_radios is None or self.rx_radios is None:
                return

            self.power_matrix=[]
            self.all_colors=[]
            
            for tx_radio in self.tx_radios:
                rx_powers = []
                rx_colors = []
                for rx_radio in self.rx_radios:
                    # powerAtRx is the same for all Rx bands, so just
                    # use the first one

                    if not (self.global_protection_level):
                        damage_threshold = self.protection_levels[rx_radio][0]
                        overload_threshold = self.protection_levels[rx_radio][1]
                        intermod_threshold = self.protection_levels[rx_radio][2]
                        desense_threshold = self.protection_levels[rx_radio][3]
                    
                    rx_band = self.rev.get_band_names(rx_radio, self.modeRx)[0]
                    if tx_radio == rx_radio:
                        # skip self-interaction
                        rx_powers.append('N/A')
                        rx_colors.append('white')
                        continue
                
                    max_power = -200
                    tx_bands = self.rev.get_band_names(tx_radio, self.modeTx)


                    for tx_band in tx_bands: 
                        # Find the highest power level at the Rx input due to each Tx Radio.
                        # Can look at any Rx freq since susceptibility won't impact
                        # powerAtRx, but need to look at all tx channels since coupling
                        # can change over a transmitter's bandwidth
                        rx_freq = self.rev.get_active_frequencies(rx_radio, rx_band, self.modeRx)[0]
                        self.domain.set_receiver(rx_radio, rx_band)            
                        self.domain.set_interferer(tx_radio, tx_band)
                        interaction = self.rev.run(self.domain)
                        self.domain.set_receiver(rx_radio, rx_band, rx_freq)
                        tx_freqs = self.rev.get_active_frequencies(tx_radio, tx_band, self.modeTx)
                        
                        power_list = []

                        for tx_freq in tx_freqs:
                            self.domain.set_interferer(tx_radio, tx_band, tx_freq)
                            instance = interaction.get_instance(self.domain)
                            power = instance.get_value(self.mode_power)

                            if power > damage_threshold:
                                classification = 'damage'
                            elif power > overload_threshold:
                                classification = 'overload'
                            elif power > intermod_threshold:
                                classification = 'intermodulation'
                            else:
                                classification = 'desensitization'

                            power_list.append(power)

                            if instance.get_value(self.mode_power) > max_power and classification in filter:
                                max_power = instance.get_value(self.mode_power)

                    # If the worst case for the band-pair is below the power thresholds, then
                    # there are no interference issues and no offset is required.
                    if max_power > -200:
                        rx_powers.append(max_power)
                        if (max_power > damage_threshold):
                            rx_colors.append('red')
                        elif (max_power > overload_threshold):
                            rx_colors.append('orange')
                        elif (max_power > intermod_threshold):
                            rx_colors.append('yellow')
                        else:
                            rx_colors.append('green')
                    else:
                        rx_powers.append("< -200")
                        rx_colors.append('white')

                self.all_colors.append(rx_colors)
                self.power_matrix.append(rx_powers)

            self.populate_table()
    
    def populate_table(self):
        ' Populate the scenario matrix table widget '
        if self.tab_widget.currentIndex() == 0:
            table = self.protection_matrix
            button = self.protection_export_btn
            img_btn = self.protection_save_img_btn
        else:
            table = self.interference_matrix
            button = self.interference_export_btn
            img_btn = self.interference_save_img_btn

        num_cols = len(self.all_colors)
        num_rows = len(self.all_colors[0])
        table.setColumnCount(num_cols)
        table.setRowCount(num_rows)
        table.setVerticalHeaderLabels(self.rx_radios)
        table.setHorizontalHeaderLabels(self.tx_radios)

        for col in range(num_cols):
            for row in range(num_rows):
                item = QtWidgets.QTableWidgetItem(str(self.power_matrix[col][row]))
                table.setItem(row, col, item)
                cell = table.item(row, col)
                cell.setBackground(self.color_dict[self.all_colors[col][row]][0])
        
        button.setEnabled(True)
        img_btn.setEnabled(True)
    
    def closeEvent(self, event):
        emitapp.close_desktop()

if __name__ == '__main__':
    app = QtWidgets.QApplication([])
    window = MainWindow()
    window.show()
    app.exec_()