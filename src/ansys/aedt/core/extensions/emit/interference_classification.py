# -*- coding: utf-8 -*-
#
# Copyright (C) 2021 - 2026 ANSYS, Inc. and/or its affiliates.
# SPDX-License-Identifier: MIT
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

from dataclasses import dataclass
import tkinter
from tkinter import messagebox
from tkinter import ttk

from ansys.aedt.core.emit_core.emit_constants import InterfererType
from ansys.aedt.core.extensions.misc import ExtensionEMITCommon
from ansys.aedt.core.extensions.misc import get_arguments

EXTENSION_TITLE = "EMIT Interference Classification"
EXTENSION_DEFAULT_ARGUMENTS = {}


@dataclass
class _MatrixData:
    tx_radios: list
    rx_radios: list
    colors: list  # colors[col][row]
    values: list  # values[col][row]


class InterferenceClassificationExtension(ExtensionEMITCommon):
    """Interactive EMIT extension for Protection Level and Interference Type classification."""

    def __init__(self, withdraw: bool = False) -> None:
        self._matrix = {"protection": None, "interference": None}
        self._filters_interf = {}
        self._filters_prot = {}
        # Color mapping canvas/excel rendering
        self._color_map = {
            "green": "#90D890",
            "yellow": "#FFEB80",
            "orange": "#FFA860",
            "red": "#FF8090",
            "white": "#FFFFFF",
        }
        self._default_protection_labels = ["Damage", "Overload", "Intermodulation", "Desensitization"]
        self._default_protection_levels = [30.0, -4.0, -30.0, -104.0]  # Damage, Overload, Intermod, Desensitization
        super().__init__(
            EXTENSION_TITLE,
            theme_color="light",
            withdraw=withdraw,
            add_custom_content=True,
        )

    def add_extension_content(self) -> None:
        root = self.root
        # Initialize Tk variables bound to the current root
        self._filters_interf = {
            "in_in": tkinter.BooleanVar(master=root, value=True),
            "out_in": tkinter.BooleanVar(master=root, value=True),
            "in_out": tkinter.BooleanVar(master=root, value=True),
            "out_out": tkinter.BooleanVar(master=root, value=True),
        }
        self._filters_prot = {
            label.lower(): tkinter.BooleanVar(master=root, value=True) for label in self._default_protection_labels
        }
        # Header with project/design info
        info = ttk.Frame(root, style="PyAEDT.TFrame")
        info.grid(row=0, column=0, sticky="ew", padx=10, pady=5)
        lbl_proj = ttk.Label(info, text=f"Project: {self.active_project_name}", style="PyAEDT.TLabel")
        lbl_proj.grid(row=0, column=0, sticky="w")
        lbl_design = ttk.Label(info, text=f"   Design: {self.active_design_name}", style="PyAEDT.TLabel")
        lbl_design.grid(row=0, column=1, sticky="w", padx=(10, 0))

        # Notebook
        nb = ttk.Notebook(root, style="TNotebook")
        nb.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)
        root.grid_rowconfigure(1, weight=1)
        root.grid_columnconfigure(0, weight=1)

        # Tabs
        prot_tab = ttk.Frame(nb, style="PyAEDT.TFrame")
        int_tab = ttk.Frame(nb, style="PyAEDT.TFrame")
        nb.add(prot_tab, text="Protection Level")
        nb.add(int_tab, text="Interference Type")

        # Protection Level tab layout
        prot_top = ttk.Frame(prot_tab, style="PyAEDT.TFrame")
        prot_top.pack(fill=tkinter.X, padx=6, pady=6)

        prot_left = ttk.LabelFrame(prot_top, text="Protection Level Thresholds", style="PyAEDT.TLabelframe")
        prot_left.pack(side=tkinter.LEFT, fill=tkinter.BOTH, expand=True, padx=(0, 6))

        for label in self._default_protection_labels:
            k = label.lower()
            ttk.Checkbutton(prot_left, text=label, variable=self._filters_prot[k], style="PyAEDT.TCheckbutton").pack(
                anchor=tkinter.W
            )

        prot_right = ttk.LabelFrame(
            prot_top, text="Protection Level Classification (Editable)", style="PyAEDT.TLabelframe"
        )
        prot_right.pack(side=tkinter.LEFT, fill=tkinter.BOTH, expand=True)

        # Canvas-based legend (more reliable color rendering than ttk.Treeview)
        legend_frame = ttk.Frame(prot_right, style="PyAEDT.TFrame")
        legend_frame.pack(fill=tkinter.X, padx=6, pady=6)

        # Header
        header_frame = ttk.Frame(legend_frame, style="PyAEDT.TFrame")
        header_frame.pack(fill=tkinter.X)
        ttk.Label(header_frame, text="Protection Level (dBm)", style="PyAEDT.TLabel").pack()

        # Legend rows with Canvas for colors
        legend_colors = [
            self._color_map["red"],
            self._color_map["orange"],
            self._color_map["yellow"],
            self._color_map["green"],
        ]
        legend_rows = [
            (label, level, color)
            for label, level, color in zip(
                self._default_protection_labels, self._default_protection_levels, legend_colors
            )
        ]

        prot_legend_entries = {}
        for label, value, color in legend_rows:
            row_frame = ttk.Frame(legend_frame, style="PyAEDT.TFrame")
            row_frame.pack(fill=tkinter.X, pady=2)

            # Color indicator (Canvas rectangle)
            canvas = tkinter.Canvas(row_frame, width=20, height=20, highlightthickness=0)
            canvas.pack(side=tkinter.LEFT, padx=(0, 6))
            canvas.create_rectangle(0, 0, 20, 20, fill=color, outline="gray")

            # Label
            ttk.Label(row_frame, text=label, width=18, style="PyAEDT.TLabel").pack(side=tkinter.LEFT)

            # Editable value entry
            entry = ttk.Entry(row_frame, width=12, style="PyAEDT.TEntry")
            entry.insert(0, value)
            entry.pack(side=tkinter.LEFT, padx=6)
            prot_legend_entries[label] = entry

        self._widgets["prot_legend_entries"] = prot_legend_entries

        # Radio-specific protection levels controls
        self._radio_specific_var = tkinter.BooleanVar(master=root, value=False)
        radio_toggle = ttk.Checkbutton(
            prot_right,
            text="Use radio specific protection levels",
            variable=self._radio_specific_var,
            command=self._on_radio_specific_toggle,
            style="PyAEDT.TCheckbutton",
        )
        radio_toggle.pack(anchor=tkinter.W, padx=6, pady=(6, 0))
        self._widgets["radio_specific_toggle"] = radio_toggle

        # Dropdown for selecting radio when radio-specific levels enabled
        self._radio_dropdown = ttk.Combobox(prot_right, state="disabled", values=[])
        self._radio_dropdown.pack(fill=tkinter.X, padx=6, pady=(4, 6))
        self._radio_dropdown.bind("<<ComboboxSelected>>", self._on_radio_dropdown_changed)
        self._widgets["radio_dropdown"] = self._radio_dropdown

        # Protection level storage
        self._global_protection_level = True
        self._protection_levels = {"Global": self._get_legend_values()}

        # Protection Level Results Matrix and Buttons
        prot_matrix = ttk.Frame(prot_tab, style="PyAEDT.TFrame")
        prot_matrix.pack(fill=tkinter.BOTH, expand=True, padx=6, pady=(0, 6))

        prot_btns = ttk.Frame(prot_tab, style="PyAEDT.TFrame")
        prot_btns.pack(fill=tkinter.X, padx=6, pady=(0, 6))
        prot_btns.grid_columnconfigure(2, weight=1)
        btn_prot_run = ttk.Button(
            prot_btns, text="Generate Results", command=self._on_run_protection, style="PyAEDT.TButton"
        )
        btn_prot_exp = ttk.Button(
            prot_btns, text="Export to Excel", command=self._on_export_excel, style="PyAEDT.TButton"
        )
        btn_prot_run.grid(row=0, column=0, padx=(0, 6), sticky="w")
        btn_prot_exp.grid(row=0, column=1, padx=(0, 6), sticky="w")

        # Add theme toggle button
        self.add_toggle_theme_button(prot_btns, toggle_row=0, toggle_column=2)

        # Interference Type tab layout
        int_top = ttk.Frame(int_tab, style="PyAEDT.TFrame")
        int_top.pack(fill=tkinter.X, padx=6, pady=6)

        int_left = ttk.LabelFrame(int_top, text="Interference Type (Source / Victim)", style="PyAEDT.TLabelframe")
        int_left.pack(side=tkinter.LEFT, fill=tkinter.BOTH, expand=True, padx=(0, 6))
        ttk.Checkbutton(
            int_left, text="Inband / Inband", variable=self._filters_interf["in_in"], style="PyAEDT.TCheckbutton"
        ).pack(anchor=tkinter.W)
        ttk.Checkbutton(
            int_left, text="Out of Band / Inband", variable=self._filters_interf["out_in"], style="PyAEDT.TCheckbutton"
        ).pack(anchor=tkinter.W)
        ttk.Checkbutton(
            int_left, text="Inband / Out of Band", variable=self._filters_interf["in_out"], style="PyAEDT.TCheckbutton"
        ).pack(anchor=tkinter.W)
        ttk.Checkbutton(
            int_left,
            text="Out of Band / Out of Band",
            variable=self._filters_interf["out_out"],
            style="PyAEDT.TCheckbutton",
        ).pack(anchor=tkinter.W)

        int_right = ttk.LabelFrame(int_top, text="Interference Type Classification", style="PyAEDT.TLabelframe")
        int_right.pack(side=tkinter.LEFT, fill=tkinter.BOTH, expand=True)

        # Canvas-based legend (more reliable color rendering than ttk.Treeview)
        int_legend_frame = ttk.Frame(int_right, style="PyAEDT.TFrame")
        int_legend_frame.pack(fill=tkinter.X, padx=6, pady=6)

        # Color-coded legend rows
        irows = [
            ("Inband / Inband", self._color_map["red"]),
            ("Out of Band / Inband", self._color_map["orange"]),
            ("Inband / Out of Band", self._color_map["yellow"]),
            ("Out of Band / Out of Band", self._color_map["green"]),
        ]

        for text, color in irows:
            row_frame = ttk.Frame(int_legend_frame, style="PyAEDT.TFrame")
            row_frame.pack(fill=tkinter.X, pady=2)

            # Color indicator (Canvas rectangle)
            canvas = tkinter.Canvas(row_frame, width=20, height=20, highlightthickness=0)
            canvas.pack(side=tkinter.LEFT, padx=(0, 6))
            canvas.create_rectangle(0, 0, 20, 20, fill=color, outline="gray")

            # Label
            ttk.Label(row_frame, text=text, style="PyAEDT.TLabel").pack(side=tkinter.LEFT)

        # Interference Type Results Matrix and Buttons
        int_matrix = ttk.Frame(int_tab, style="PyAEDT.TFrame")
        int_matrix.pack(fill=tkinter.BOTH, expand=True, padx=6, pady=(0, 6))

        int_btns = ttk.Frame(int_tab, style="PyAEDT.TFrame")
        int_btns.pack(fill=tkinter.X, padx=6, pady=(0, 6))
        int_btns.grid_columnconfigure(2, weight=1)
        btn_int_run = ttk.Button(
            int_btns, text="Generate Results", command=self._on_run_interference, style="PyAEDT.TButton"
        )
        btn_int_exp = ttk.Button(
            int_btns, text="Export to Excel", command=self._on_export_excel, style="PyAEDT.TButton"
        )
        btn_int_run.grid(row=0, column=0, padx=(0, 6), sticky="w")
        btn_int_exp.grid(row=0, column=1, padx=(0, 6), sticky="w")

        # Add theme toggle button
        self.add_toggle_theme_button(int_btns, toggle_row=0, toggle_column=2)

        # Matrix canvases: one per tab, color-coded like the Tk example
        canvas_prot = tkinter.Canvas(prot_matrix, highlightthickness=0, background="white")
        canvas_int = tkinter.Canvas(int_matrix, highlightthickness=0, background="white")
        canvas_prot.pack(fill=tkinter.BOTH, expand=True, padx=6, pady=(0, 6))
        canvas_int.pack(fill=tkinter.BOTH, expand=True, padx=6, pady=(0, 6))
        self._widgets["canvas_prot"] = canvas_prot
        self._widgets["canvas_int"] = canvas_int

    # ---------------- Event handlers -----------------
    def _on_run_interference(self) -> None:
        """Generate interference type classification results."""
        try:
            filters = self._build_interf_filters()
            tx_radios, rx_radios, colors, values = self._compute_interference(filters)
            self._matrix["interference"] = _MatrixData(tx_radios, rx_radios, colors, values)
            self._render_matrix(tab="interference")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate interference results: {e}")

    def _on_run_protection(self) -> None:
        """Generate protection level classification results."""
        try:
            filters = [k for k, v in self._filters_prot.items() if bool(v.get())]
            tx_radios, rx_radios, colors, values = self._compute_protection(filters)
            self._matrix["protection"] = _MatrixData(tx_radios, rx_radios, colors, values)
            self._render_matrix(tab="protection")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate protection results: {e}")

    # ---------------- Legend/radio helpers -----------------
    def _get_legend_values(self):
        """Retrieve protection level values from the legend entry widgets."""
        vals = []
        entries = self._widgets.get("prot_legend_entries", {})
        for label in self._default_protection_labels:
            entry = entries.get(label)
            if entry:
                try:
                    vals.append(float(entry.get()))
                except Exception:
                    vals.append(0.0)
            else:
                vals.append(0.0)
        return vals

    def _on_radio_specific_toggle(self) -> None:
        """Configure legend and protection levels when radio-specific toggle changes."""
        enabled = bool(self._radio_specific_var.get())
        if enabled:
            # Build per-radio levels from current legend values
            app = self.aedt_application
            values = self._get_legend_values()
            try:
                rx_names = app.results.analyze().get_receiver_names()
            except Exception:
                rx_names = []
            self._protection_levels = {name: values[:] for name in rx_names}
            self._radio_dropdown.configure(state="readonly", values=rx_names)
            if rx_names:
                self._radio_dropdown.set(rx_names[0])
                self._current_radio = rx_names[0]  # Track initial radio
            self._global_protection_level = False
        else:
            values = self._get_legend_values()
            self._protection_levels = {"Global": values}
            self._radio_dropdown.configure(state="disabled", values=[])
            self._radio_dropdown.set("")
            self._current_radio = None  # Clear tracking when back to global
            self._global_protection_level = True

    def _on_radio_dropdown_changed(self, _evt=None) -> None:
        """Update legend entry values when selected radio changes."""
        # First, save the current Entry values to the previously selected radio
        prev_radio = getattr(self, "_current_radio", None)
        if prev_radio and prev_radio in self._protection_levels:
            current_values = self._get_legend_values()
            self._protection_levels[prev_radio] = current_values

        # Now load the newly selected radio's values
        cur = self._radio_dropdown.get()
        if not cur:
            return
        if cur not in self._protection_levels:
            return

        # Remember this as the current radio for next switch
        self._current_radio = cur

        values = self._protection_levels[cur]
        entries = self._widgets.get("prot_legend_entries", {})

        for i, label in enumerate(self._default_protection_labels):
            entry = entries.get(label)
            if entry and i < len(values):
                entry.delete(0, tkinter.END)
                entry.insert(0, str(values[i]))

    def _on_export_excel(self) -> None:
        """Export the current results matrix to an Excel file."""
        from tkinter import filedialog

        import openpyxl
        from openpyxl.styles import PatternFill

        # Determine which tab is active
        notebook = self.root.nametowidget(".!notebook")
        current_tab_index = notebook.index(notebook.select())
        tab_key = "protection" if current_tab_index == 0 else "interference"
        matrix = self._matrix.get(tab_key)

        if not matrix:
            messagebox.showwarning("No data", "Please generate results first.")
            return
        default_name = "Interference Classification"  # default; user can override
        fname = filedialog.asksaveasfilename(
            title="Save Scenario Matrix",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("xlsx", "*.xlsx")],
        )
        if not fname:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        header = ["Tx/Rx"] + matrix.tx_radios
        ws.append(header)
        # rows
        for r, rx in enumerate(matrix.rx_radios):
            row = [rx]
            for c, _tx in enumerate(matrix.tx_radios):
                val = matrix.values[c][r] if matrix.values else ""
                row.append(val)
            ws.append(row)
        # apply fills using color map
        for c in range(len(matrix.tx_radios)):
            for r in range(len(matrix.rx_radios)):
                raw = matrix.colors[c][r]
                hexcol = self._color_map.get(str(raw).lower(), "#FFFFFF").lstrip("#")
                cell = ws.cell(row=2 + r, column=2 + c)
                cell.fill = PatternFill(start_color=hexcol, end_color=hexcol, fill_type="solid")
        wb.save(fname)

    # --------------- Computation helpers ----------------
    def _build_interf_filters(self):
        """Build list of EMIT filter strings based on UI selections."""
        # Maps to EMIT filter strings used by Results APIs
        all_filters = [
            "TxFundamental:In-band",
            ["TxHarmonic/Spurious:In-band", "Intermod:In-band", "Broadband:In-band"],
            "TxFundamental:Out-of-band",
            ["TxHarmonic/Spurious:Out-of-band", "Intermod:Out-of-band", "Broadband:Out-of-band"],
        ]
        checks = [
            bool(self._filters_interf["in_in"].get()),
            bool(self._filters_interf["out_in"].get()),
            bool(self._filters_interf["in_out"].get()),
            bool(self._filters_interf["out_out"].get()),
        ]
        return [f for f, v in zip(all_filters, checks) if v]

    def _compute_interference(self, filter_list):
        """Compute interference type classification matrix."""
        app = self.aedt_application
        # Radios
        try:
            radios = app.modeler.components.get_radios()
        except Exception:  # pragma: no cover
            radios = []
        if len(radios) < 2:
            raise RuntimeError("At least two radios are required.")

        domain = app.results.interaction_domain()
        # Prefer API on results viewer if available; otherwise fallback to results
        rev = app.results.analyze()
        colors, matrix = rev.interference_type_classification(
            domain, interferer_type=InterfererType().TRANSMITTERS, use_filter=True, filter_list=filter_list
        )
        tx = rev.get_interferer_names(InterfererType().TRANSMITTERS_AND_EMITTERS)
        rx = rev.get_receiver_names()
        return tx, rx, colors, matrix

    def _compute_protection(self, filter_list):
        """Compute protection level classification matrix."""
        app = self.aedt_application
        try:
            radios = app.modeler.components.get_radios()
        except Exception:
            radios = []
        if len(radios) < 2:
            raise RuntimeError("At least two radios are required.")

        # Always get current values from Entry widgets to capture any edits
        current_values = self._get_legend_values()

        # Update stored values
        if self._global_protection_level:
            self._protection_levels["Global"] = current_values
            global_levels = current_values
        else:
            # Update the currently selected radio's values
            cur_radio = self._radio_dropdown.get()
            if cur_radio:
                self._protection_levels[cur_radio] = current_values
            global_levels = self._protection_levels.get("Global", current_values)

        rev = app.results.analyze()
        domain = app.results.interaction_domain()

        colors, matrix = rev.protection_level_classification(
            domain=domain,
            interferer_type=InterfererType().TRANSMITTERS,
            global_protection_level=self._global_protection_level,
            global_levels=global_levels,
            protection_levels=self._protection_levels,
            use_filter=True,
            filter_list=filter_list,
        )

        tx = rev.get_interferer_names(InterfererType().TRANSMITTERS)
        rx = rev.get_receiver_names()
        return tx, rx, colors, matrix

    # --------------- UI rendering helpers ----------------
    def _render_matrix(self, tab: str) -> None:
        """Display the results matrix on the canvas for the given tab."""
        matrix = self._matrix.get(tab)
        if not matrix:
            return
        # Choose the correct canvas for the tab
        cnv = self._widgets["canvas_int"] if tab == "interference" else self._widgets["canvas_prot"]

        # Draw a resizable grid with per-cell backgrounds and values
        def draw_table(_event=None) -> None:
            cnv.delete("all")
            W = max(cnv.winfo_width(), 200)
            H = max(cnv.winfo_height(), 150)
            margin_left = 120
            margin_top = 26
            grid_x0 = 1
            grid_y0 = 1
            grid_x1 = W - 2
            grid_y1 = H - 2

            header_w = margin_left
            header_h = margin_top
            cell_x0 = grid_x0 + header_w
            cell_y0 = grid_y0 + header_h
            cell_w = max(grid_x1 - cell_x0, 50)
            cell_h = max(grid_y1 - cell_y0, 50)

            num_cols = len(matrix.tx_radios)
            num_rows = len(matrix.rx_radios)
            if num_cols == 0 or num_rows == 0:
                return

            col_w = cell_w / max(num_cols, 1)
            row_h = cell_h / max(num_rows, 1)

            # Row headers (Rx)
            for r in range(num_rows):
                y0 = cell_y0 + r * row_h
                y1 = y0 + row_h
                cnv.create_rectangle(grid_x0, y0, grid_x0 + header_w, y1, fill="#f2f2f2", outline="#cccccc")
                cnv.create_text(grid_x0 + 6, (y0 + y1) / 2, text=str(matrix.rx_radios[r]), anchor="w")

            # Column headers (Tx)
            for c in range(num_cols):
                x0 = cell_x0 + c * col_w
                x1 = x0 + col_w
                cnv.create_rectangle(x0, grid_y0, x1, grid_y0 + header_h, fill="#f2f2f2", outline="#cccccc")
                cnv.create_text((x0 + x1) / 2, grid_y0 + header_h / 2, text=str(matrix.tx_radios[c]), anchor="center")

            # Cells
            for r in range(num_rows):
                for c in range(num_cols):
                    x0 = cell_x0 + c * col_w
                    y0 = cell_y0 + r * row_h
                    x1 = x0 + col_w
                    y1 = y0 + row_h
                    # Map EMIT color name to hex
                    try:
                        raw = matrix.colors[c][r]
                    except Exception:
                        raw = None
                    hexcol = self._color_map.get(str(raw).lower(), "#FFFFFF")
                    cnv.create_rectangle(x0, y0, x1, y1, fill=hexcol, outline="#ffffff")
                    # Value text
                    try:
                        val = matrix.values[c][r]
                    except Exception:
                        val = ""
                    cnv.create_text((x0 + x1) / 2, (y0 + y1) / 2, text=str(val), anchor="center")

            cnv.update_idletasks()

        cnv.bind("<Configure>", draw_table)
        draw_table()


if __name__ == "__main__":  # pragma: no cover
    args = get_arguments(EXTENSION_DEFAULT_ARGUMENTS, EXTENSION_TITLE)
    ext = InterferenceClassificationExtension(withdraw=False)
    tkinter.mainloop()
